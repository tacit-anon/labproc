#!/usr/bin/env python3
"""
Compile labeled annotations into the standard Tacit bundle.

Input JSON shape:
[
  {
    "video_file": "good_yt_xyz.mp4",
    "branch": "op",
    "category": "organic_purification",
    "timestamp_seconds": 150.5,
    "label": "crystals_forming",
    "confidence": "high",
    "frame_path": "/abs/path/to/extracted/frame.jpg",
    "folder": "Videos/Organic Purification"
  },
  ...
]

Output bundle:
  {output_root}/
    {video_basename}/
      t{HHHHH.SS}__{label}.jpg
    tacit_annotations_{YYYY-MM-DD}.xlsx

Usage:
  build_bundle.py <annotations.json> <output_root>

Requires: openpyxl  (pip install openpyxl --break-system-packages)
"""
import json, sys, os, shutil, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


HEADERS = ['branch', 'video_file', 'timestamp_seconds', 'physical_state',
           'confidence', 'your_label', 'screenshot_path']
MANIFEST_HEADERS = ['video_file', 'category', 'branch', 'status',
                    'annotation_count', 'folder', 'last_modified']


def basename_no_ext(name):
    base = os.path.splitext(os.path.basename(name))[0]
    return ''.join(c if c.isalnum() or c in '_-' else '_' for c in base)


def safe_fs(s):
    return ''.join(c if c.isalnum() or c in '_-' else '_' for c in str(s))


def padded_ts(ts):
    int_part, dec_part = f'{float(ts):.2f}'.split('.')
    return f'{int(int_part):05d}.{dec_part}'


def shot_filename(ts, label):
    return f't{padded_ts(ts)}__{safe_fs(label)}.jpg'


def write_xlsx(rows, videos, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'annotations'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4F46E5')

    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='left')

    for r, row in enumerate(rows, start=2):
        for i, h in enumerate(HEADERS, start=1):
            ws.cell(row=r, column=i, value=row[h])

    for i, h in enumerate(HEADERS, start=1):
        ws.column_dimensions[chr(64 + i)].width = max(len(h) + 2, 22)

    ws2 = wb.create_sheet('video_manifest')
    for i, h in enumerate(MANIFEST_HEADERS, start=1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill

    for r, (vf, info) in enumerate(videos.items(), start=2):
        ws2.cell(row=r, column=1, value=vf)
        ws2.cell(row=r, column=2, value=info.get('category', ''))
        ws2.cell(row=r, column=3, value=info.get('branch', ''))
        ws2.cell(row=r, column=4, value='in_progress')
        ws2.cell(row=r, column=5, value=info['count'])
        ws2.cell(row=r, column=6, value=info.get('folder', ''))
        ws2.cell(row=r, column=7, value=datetime.datetime.now().isoformat(timespec='seconds'))

    for i, h in enumerate(MANIFEST_HEADERS, start=1):
        ws2.column_dimensions[chr(64 + i)].width = max(len(h) + 2, 18)

    wb.save(out_path)


def main():
    if len(sys.argv) < 3:
        print('Usage: build_bundle.py <annotations.json> <output_root>', file=sys.stderr)
        sys.exit(1)

    annos_path = sys.argv[1]
    out_root = sys.argv[2]
    os.makedirs(out_root, exist_ok=True)

    with open(annos_path) as f:
        annos = json.load(f)

    rows = []
    videos = {}

    for a in annos:
        video_file = a['video_file']
        video_base = basename_no_ext(video_file)
        ts = float(a['timestamp_seconds'])
        label = a['label']
        branch = a['branch']

        video_dir = os.path.join(out_root, video_base)
        os.makedirs(video_dir, exist_ok=True)
        fname = shot_filename(ts, label)
        dst = os.path.join(video_dir, fname)
        shutil.copy(a['frame_path'], dst)

        rel = f'{video_base}/{fname}'
        rows.append({
            'branch': branch,
            'video_file': video_file,
            'timestamp_seconds': ts,
            'physical_state': label,
            'confidence': a.get('confidence', 'medium'),
            'your_label': label,
            'screenshot_path': rel,
        })

        info = videos.setdefault(video_file, {
            'category': a.get('category', ''),
            'branch': branch,
            'folder': a.get('folder', ''),
            'count': 0,
        })
        info['count'] += 1

    date_str = datetime.date.today().isoformat()
    xlsx_path = os.path.join(out_root, f'tacit_annotations_{date_str}.xlsx')
    write_xlsx(rows, videos, xlsx_path)

    print(f'Wrote {len(rows)} annotations across {len(videos)} videos.')
    print(f'  Bundle:      {out_root}/')
    print(f'  Spreadsheet: {xlsx_path}')


if __name__ == '__main__':
    main()
