#!/usr/bin/env python3
"""
Append a per-batch annotations.json (or a built bundle's xlsx) into the
master corpus xlsx.

Enforces:
  - 10-column annotations sheet schema
  - dedup against (branch, video_file, timestamp_seconds) triple
  - video_manifest sheet (per-video status / count / folder / last_modified)
  - video_audit sheet (NON_CATEGORY_SKIP / TAXONOMY_GAP entries)
  - per-branch CSV exports next to the master xlsx (one CSV per branch)

Usage:
  python append_to_master.py <annotations.json> <master.xlsx>
  python append_to_master.py <bundle_xlsx> <master.xlsx>

Idempotent: re-running with the same input adds zero rows.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections.abc import Iterable
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    print("error: openpyxl required. install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


ANNOTATIONS_COLUMNS = [
    "branch",
    "video_file",
    "timestamp_seconds",
    "physical_state",
    "confidence",
    "your_label",
    "screenshot_path",
    "substance_tags",
    "action_tags",
    "equipment_tags",
]

MANIFEST_COLUMNS = [
    "video_file",
    "category",
    "branch",
    "status",
    "annotation_count",
    "folder",
    "last_modified",
]

AUDIT_COLUMNS = [
    "video_file",
    "branch",
    "decision",
    "reason",
    "label_proposal",
    "logged_at",
]

BRANCH_TO_CATEGORY = {
    "op": "organic_purification",
    "pcr": "pcr",
    "wb": "western_blot",
}


def utcnow_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def load_input_rows(path: Path) -> tuple[list[dict], list[dict]]:
    """
    Returns (annotation_rows, audit_rows).
    Audit rows come from a top-level "audit" key in JSON, or empty for xlsx.
    """
    if path.suffix == ".json":
        data = json.loads(path.read_text())
        branch = data.get("branch")
        ann_rows = []
        for a in data.get("annotations", []):
            row = {
                "branch": a.get("branch", branch),
                "video_file": a["video_file"],
                "timestamp_seconds": float(a["ts"] if "ts" in a else a["timestamp_seconds"]),
                "physical_state": a.get("physical_state", a.get("label", a.get("your_label", ""))),
                "confidence": a.get("confidence", "medium"),
                "your_label": a.get("your_label", a.get("label", "")),
                "screenshot_path": a.get("screenshot_path", ""),
                "substance_tags": a.get("substance_tags", ""),
                "action_tags": a.get("action_tags", ""),
                "equipment_tags": a.get("equipment_tags", ""),
            }
            ann_rows.append(row)
        audit_rows = data.get("audit", [])
        return ann_rows, audit_rows

    if path.suffix in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["annotations"] if "annotations" in wb.sheetnames else wb.active
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        idx = {col: header.index(col) for col in ANNOTATIONS_COLUMNS if col in header}
        ann_rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = {col: (r[idx[col]] if col in idx else "") for col in ANNOTATIONS_COLUMNS}
            if row["video_file"]:
                ann_rows.append(row)
        return ann_rows, []

    raise ValueError(f"unsupported input extension: {path.suffix}")


def open_or_create_master(master_path: Path) -> Workbook:
    if master_path.exists():
        return openpyxl.load_workbook(master_path)
    wb = Workbook()
    default = wb.active
    default.title = "annotations"
    default.append(ANNOTATIONS_COLUMNS)
    manifest = wb.create_sheet("video_manifest")
    manifest.append(MANIFEST_COLUMNS)
    audit = wb.create_sheet("video_audit")
    audit.append(AUDIT_COLUMNS)
    return wb


def existing_keys(ws) -> set[tuple[str, str, float]]:
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    i_branch = header.index("branch")
    i_video = header.index("video_file")
    i_ts = header.index("timestamp_seconds")
    keys = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[i_video]:
            keys.add((r[i_branch], r[i_video], float(r[i_ts])))
    return keys


def upsert_manifest(wb: Workbook, ann_rows: Iterable[dict], folder_hint: str) -> None:
    manifest = wb["video_manifest"] if "video_manifest" in wb.sheetnames else wb.create_sheet("video_manifest")
    if manifest.max_row == 0:
        manifest.append(MANIFEST_COLUMNS)
    header = [c.value for c in next(manifest.iter_rows(max_row=1))]
    i_video = header.index("video_file")
    existing = {row[i_video]: idx + 2 for idx, row in enumerate(manifest.iter_rows(min_row=2, values_only=True)) if row[i_video]}

    counts: dict[tuple[str, str], int] = {}
    for r in ann_rows:
        counts[(r["video_file"], r["branch"])] = counts.get((r["video_file"], r["branch"]), 0) + 1

    for (video_file, branch), count in counts.items():
        category = BRANCH_TO_CATEGORY.get(branch, branch)
        ts = utcnow_iso()
        if video_file in existing:
            row_idx = existing[video_file]
            manifest.cell(row=row_idx, column=4, value="done")
            current = manifest.cell(row=row_idx, column=5).value or 0
            manifest.cell(row=row_idx, column=5, value=int(current) + count)
            manifest.cell(row=row_idx, column=7, value=ts)
        else:
            manifest.append([video_file, category, branch, "done", count, folder_hint, ts])


def append_audit(wb: Workbook, audit_rows: Iterable[dict]) -> None:
    """Append audit rows, dedup on (video_file, branch, decision, reason)."""
    if not audit_rows:
        return
    audit = wb["video_audit"] if "video_audit" in wb.sheetnames else wb.create_sheet("video_audit")
    if audit.max_row == 0:
        audit.append(AUDIT_COLUMNS)
    existing_keys = set()
    for r in audit.iter_rows(min_row=2, values_only=True):
        if r[0]:
            existing_keys.add((r[0], r[1], r[2], r[3]))
    ts = utcnow_iso()
    for r in audit_rows:
        key = (r.get("video_file", ""), r.get("branch", ""), r.get("decision", ""), r.get("reason", ""))
        if key in existing_keys:
            continue
        existing_keys.add(key)
        audit.append([
            r.get("video_file", ""),
            r.get("branch", ""),
            r.get("decision", ""),
            r.get("reason", ""),
            r.get("label_proposal", ""),
            r.get("logged_at", ts),
        ])


def write_branch_csvs(master_path: Path) -> dict[str, int]:
    """Regenerate per-branch CSV files next to master.xlsx."""
    wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
    ws = wb["annotations"]
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    i_branch = header.index("branch")
    i_video = header.index("video_file")
    i_ts = header.index("timestamp_seconds")

    by_branch: dict[str, list[list]] = {}
    seen_videos: dict[str, dict[str, int]] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[i_video]:
            continue
        branch = r[i_branch]
        seen_videos.setdefault(branch, {})
        if r[i_video] not in seen_videos[branch]:
            seen_videos[branch][r[i_video]] = len(seen_videos[branch])
        vid_idx = seen_videos[branch][r[i_video]]
        clip_id = f"{branch}_{vid_idx:03d}_{int(r[i_ts])}s"
        by_branch.setdefault(branch, []).append([clip_id] + list(r))

    counts = {}
    for branch, rows in by_branch.items():
        out_path = master_path.with_name(f"{master_path.stem}_{branch}.csv")
        with out_path.open("w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["clip_id"] + header)
            w.writerows(rows)
        counts[branch] = len(rows)
    return counts


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("input_path", type=Path, help="annotations.json or bundle xlsx")
    p.add_argument("master_path", type=Path, help="master corpus xlsx")
    p.add_argument("--folder", default="", help="source folder hint for manifest")
    args = p.parse_args()

    if not args.input_path.exists():
        print(f"error: input not found: {args.input_path}", file=sys.stderr)
        return 2

    ann_rows, audit_rows = load_input_rows(args.input_path)
    print(f"loaded {len(ann_rows)} annotation rows, {len(audit_rows)} audit rows")

    wb = open_or_create_master(args.master_path)
    ws = wb["annotations"]
    keys = existing_keys(ws)

    appended = 0
    skipped_dup = 0
    for r in ann_rows:
        key = (r["branch"], r["video_file"], float(r["timestamp_seconds"]))
        if key in keys:
            skipped_dup += 1
            continue
        ws.append([r[c] for c in ANNOTATIONS_COLUMNS])
        keys.add(key)
        appended += 1

    upsert_manifest(wb, [r for r in ann_rows if (r["branch"], r["video_file"], float(r["timestamp_seconds"])) in keys], args.folder)
    append_audit(wb, audit_rows)

    wb.save(args.master_path)

    csv_counts = write_branch_csvs(args.master_path)

    print(f"appended {appended} rows, skipped {skipped_dup} duplicates")
    print(f"per-branch CSV exports: {csv_counts}")
    print(f"master saved: {args.master_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
