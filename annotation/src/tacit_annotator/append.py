"""
Append a per-batch annotations.json (or built-bundle xlsx) into the master corpus xlsx.

Enforces:
  - 10-column annotations sheet schema
  - dedupe on (branch, video_file, timestamp_seconds)
  - video_manifest upsert
  - video_audit dedupe on (video_file, branch, decision, reason)
  - per-branch CSV regeneration (one CSV next to master xlsx)
"""

from __future__ import annotations

import csv
import datetime as _dt
import json
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl import Workbook

from .schema import (
    ANNOTATION_COLUMNS,
    AUDIT_COLUMNS,
    BRANCH_TO_CATEGORY,
    MANIFEST_COLUMNS,
)


def _utcnow_iso() -> str:
    return _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


@dataclass
class AppendStats:
    appended: int
    skipped_duplicates: int
    audit_added: int
    audit_skipped_duplicates: int
    csv_counts: dict[str, int]


def _open_or_create_master(master_path: Path) -> Workbook:
    if master_path.exists():
        return openpyxl.load_workbook(master_path)
    wb = Workbook()
    default = wb.active
    default.title = "annotations"
    default.append(ANNOTATION_COLUMNS)
    manifest = wb.create_sheet("video_manifest")
    manifest.append(MANIFEST_COLUMNS)
    audit = wb.create_sheet("video_audit")
    audit.append(AUDIT_COLUMNS)
    return wb


def _load_input(path: Path) -> tuple[list[dict], list[dict]]:
    """Returns (annotation_rows, audit_rows). Accepts both .json and .xlsx inputs."""
    if path.suffix == ".json":
        data = json.loads(path.read_text())
        branch = data.get("branch", "")
        ann_rows = []
        for a in data.get("annotations", []):
            ann_rows.append({
                "branch": a.get("branch", branch),
                "video_file": a["video_file"],
                "timestamp_seconds": float(a.get("timestamp_seconds", a.get("ts", 0))),
                "physical_state": a.get("physical_state", a.get("your_label", a.get("label", ""))),
                "confidence": a.get("confidence", "medium"),
                "your_label": a.get("your_label", a.get("label", "")),
                "screenshot_path": a.get("screenshot_path", ""),
                "substance_tags": a.get("substance_tags", ""),
                "action_tags": a.get("action_tags", ""),
                "equipment_tags": a.get("equipment_tags", ""),
            })
        return ann_rows, data.get("audit", [])

    if path.suffix in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["annotations"] if "annotations" in wb.sheetnames else wb.active
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        idx = {col: header.index(col) for col in ANNOTATION_COLUMNS if col in header}
        ann_rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = {col: (r[idx[col]] if col in idx else "") for col in ANNOTATION_COLUMNS}
            if row["video_file"]:
                ann_rows.append(row)
        return ann_rows, []

    raise ValueError(f"unsupported input extension: {path.suffix}")


def _existing_annotation_keys(ws) -> set[tuple[str, str, float]]:
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    i_branch = header.index("branch")
    i_video = header.index("video_file")
    i_ts = header.index("timestamp_seconds")
    keys: set[tuple[str, str, float]] = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[i_video]:
            keys.add((r[i_branch], r[i_video], float(r[i_ts])))
    return keys


def _upsert_manifest(wb: Workbook, ann_rows: list[dict], folder_hint: str) -> None:
    manifest = wb["video_manifest"] if "video_manifest" in wb.sheetnames else wb.create_sheet("video_manifest")
    if manifest.max_row == 0:
        manifest.append(MANIFEST_COLUMNS)

    header = [c.value for c in next(manifest.iter_rows(max_row=1))]
    i_video = header.index("video_file")
    existing = {
        row[i_video]: idx + 2
        for idx, row in enumerate(manifest.iter_rows(min_row=2, values_only=True))
        if row[i_video]
    }

    counts: dict[tuple[str, str], int] = {}
    for r in ann_rows:
        counts[(r["video_file"], r["branch"])] = counts.get((r["video_file"], r["branch"]), 0) + 1

    ts = _utcnow_iso()
    for (video_file, branch), count in counts.items():
        category = BRANCH_TO_CATEGORY.get(branch, branch)
        if video_file in existing:
            row_idx = existing[video_file]
            manifest.cell(row=row_idx, column=4, value="done")
            current = manifest.cell(row=row_idx, column=5).value or 0
            manifest.cell(row=row_idx, column=5, value=int(current) + count)
            manifest.cell(row=row_idx, column=7, value=ts)
        else:
            manifest.append([video_file, category, branch, "done", count, folder_hint, ts])


def _append_audit(wb: Workbook, audit_rows: list[dict]) -> tuple[int, int]:
    audit = wb["video_audit"] if "video_audit" in wb.sheetnames else wb.create_sheet("video_audit")
    if audit.max_row == 0:
        audit.append(AUDIT_COLUMNS)

    existing_keys = set()
    for r in audit.iter_rows(min_row=2, values_only=True):
        if r[0]:
            existing_keys.add((r[0], r[1], r[2], r[3]))

    added = 0
    skipped = 0
    ts = _utcnow_iso()
    for r in audit_rows:
        key = (
            r.get("video_file", ""),
            r.get("branch", ""),
            r.get("decision", ""),
            r.get("reason", ""),
        )
        if key in existing_keys:
            skipped += 1
            continue
        existing_keys.add(key)
        added += 1
        audit.append([
            r.get("video_file", ""),
            r.get("branch", ""),
            r.get("decision", ""),
            r.get("reason", ""),
            r.get("label_proposal", ""),
            r.get("logged_at", ts),
        ])
    return added, skipped


def _write_branch_csvs(master_path: Path) -> dict[str, int]:
    """Regenerate per-branch CSV files next to master.xlsx — clip_id-keyed."""
    wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
    ws = wb["annotations"]
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    i_branch = header.index("branch")
    i_video = header.index("video_file")
    i_ts = header.index("timestamp_seconds")

    by_branch: dict[str, list[list]] = {}
    seen_videos: dict[str, dict[str, int]] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[i_video]:
            continue
        branch = r[i_branch]
        seen_videos.setdefault(branch, {})
        if r[i_video] not in seen_videos[branch]:
            seen_videos[branch][r[i_video]] = len(seen_videos[branch])
        vid_idx = seen_videos[branch][r[i_video]]
        clip_id = f"{branch}_{vid_idx:03d}_{int(r[i_ts])}s"
        by_branch.setdefault(branch, []).append([clip_id, *list(r)])

    counts = {}
    for branch, rows in by_branch.items():
        out_path = master_path.with_name(f"{master_path.stem}_{branch}.csv")
        with out_path.open("w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["clip_id", *header])
            w.writerows(rows)
        counts[branch] = len(rows)
    return counts


def append_to_master(
    input_path: Path,
    master_path: Path,
    folder_hint: str = "",
) -> AppendStats:
    """
    Idempotent append. Returns counts of appended / skipped rows.
    """
    if not input_path.exists():
        raise FileNotFoundError(input_path)

    ann_rows, audit_rows = _load_input(input_path)

    wb = _open_or_create_master(master_path)
    ws = wb["annotations"]
    keys = _existing_annotation_keys(ws)

    appended = 0
    skipped_dup = 0
    for r in ann_rows:
        key = (r["branch"], r["video_file"], float(r["timestamp_seconds"]))
        if key in keys:
            skipped_dup += 1
            continue
        ws.append([r[c] for c in ANNOTATION_COLUMNS])
        keys.add(key)
        appended += 1

    new_rows = [r for r in ann_rows if (r["branch"], r["video_file"], float(r["timestamp_seconds"])) in keys]
    _upsert_manifest(wb, new_rows, folder_hint)

    audit_added, audit_skipped = _append_audit(wb, audit_rows)

    wb.save(master_path)
    csv_counts = _write_branch_csvs(master_path)

    return AppendStats(
        appended=appended,
        skipped_duplicates=skipped_dup,
        audit_added=audit_added,
        audit_skipped_duplicates=audit_skipped,
        csv_counts=csv_counts,
    )
